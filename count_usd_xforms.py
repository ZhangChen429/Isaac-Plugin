import argparse
import csv
import os
import re
import sys
from pathlib import Path


ISAAC_ROOT = Path(r"E:\SoftAPP\isaac-sim-standalone-5.1.0-windows-x86_64")
USD_LIB_EXT = ISAAC_ROOT / "extscache" / "omni.usd.libs-1.0.1+69cbf6ad.wx64.r.cp311"
CODEX_PYTHON_SITE = Path(
    r"C:\Users\Administrator\.cache\codex-runtimes\codex-primary-runtime\dependencies\python\Lib\site-packages"
)
DEFAULT_SOURCE = Path(r"E:\Data\USD\kook")
DEFAULT_OUTPUT = Path(r"E:\Data\USD\procedure\usd_xform_counts.csv")
USD_EXTENSIONS = {".usd", ".usda", ".usdc", ".usdz"}


def setup_usd_runtime():
    sys.path.insert(0, str(USD_LIB_EXT))
    if CODEX_PYTHON_SITE.exists():
        sys.path.append(str(CODEX_PYTHON_SITE))
    dll_dirs = [
        ISAAC_ROOT,
        ISAAC_ROOT / "kit",
        ISAAC_ROOT / "kit" / "plugins",
        ISAAC_ROOT / "kit" / "kernel" / "plugins",
        ISAAC_ROOT / "kit" / "plugins" / "bindings-python",
        USD_LIB_EXT / "bin",
    ]
    for dll_dir in dll_dirs:
        if dll_dir.exists():
            os.add_dll_directory(str(dll_dir))


def natural_key(path: Path, base: Path):
    text = str(path.relative_to(base))
    return [
        int(part) if part.isdigit() else part.casefold()
        for part in re.split(r"(\d+)", text)
    ]


def parse_args():
    parser = argparse.ArgumentParser(description="Count and validate USD files.")
    parser.add_argument(
        "source",
        nargs="?",
        default=str(DEFAULT_SOURCE),
        help=f"Source file or folder. Default: {DEFAULT_SOURCE}",
    )
    parser.add_argument(
        "-o",
        "--output",
        default=str(DEFAULT_OUTPUT),
        help=f"Output CSV path. Default: {DEFAULT_OUTPUT}",
    )
    parser.add_argument(
        "--no-fix",
        action="store_true",
        help="Only read workbook counts; do not update mismatched D2 joint counts.",
    )
    return parser.parse_args()


def collect_usd_files(source: Path):
    if source.is_file():
        return [source] if source.suffix.casefold() in USD_EXTENSIONS else []
    files = [
        path
        for path in source.rglob("*")
        if path.is_file() and path.suffix.casefold() in USD_EXTENSIONS
    ]
    return sorted(files, key=lambda path: natural_key(path, source))


def get_root_xform(stage, usd_geom):
    default_prim = stage.GetDefaultPrim()
    if default_prim and default_prim.IsValid() and default_prim.IsA(usd_geom.Xform):
        return default_prim
    for prim in stage.GetPseudoRoot().GetChildren():
        if prim.IsA(usd_geom.Xform):
            return prim
    for prim in stage.TraverseAll():
        if prim.IsA(usd_geom.Xform):
            return prim
    return None


def has_articulation_root_api(prim, usd_physics):
    if not prim or not prim.IsValid():
        return False
    return prim.HasAPI(usd_physics.ArticulationRootAPI)


def inspect_usd(usd_path: Path):
    from pxr import Usd, UsdGeom, UsdPhysics

    stage = Usd.Stage.Open(str(usd_path))
    if stage is None:
        raise RuntimeError("Usd.Stage.Open returned None")

    root_xform = get_root_xform(stage, UsdGeom)
    root_path = root_xform.GetPath() if root_xform else None
    root_has_articulation = has_articulation_root_api(root_xform, UsdPhysics)

    xform_total = 0
    group_named = 0
    non_group_non_root = 0
    joint_total = 0
    articulation_paths = []

    for prim in stage.TraverseAll():
        if prim.IsA(UsdGeom.Xform):
            xform_total += 1
            if prim.GetName().startswith("group"):
                group_named += 1
            elif root_path is None or prim.GetPath() != root_path:
                non_group_non_root += 1

        if prim.IsA(UsdPhysics.Joint):
            joint_total += 1

        if has_articulation_root_api(prim, UsdPhysics):
            articulation_paths.append(str(prim.GetPath()))

    if root_has_articulation:
        articulation_status = "OK"
    elif articulation_paths:
        articulation_status = "ArticulationRoot不在根Xform"
    else:
        articulation_status = "缺少ArticulationRoot"

    return {
        "root_name": root_xform.GetName() if root_xform else "",
        "root_path": str(root_path) if root_path else "",
        "root_has_articulation": "是" if root_has_articulation else "否",
        "articulation_paths": ";".join(articulation_paths),
        "articulation_status": articulation_status,
        "xform_total": xform_total,
        "group_named": group_named,
        "non_group_non_root": non_group_non_root,
        "joint_total": joint_total,
    }


def as_int(value):
    if value is None or value == "":
        return None
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def read_and_fix_delivery_counts(folder: Path, usd_joint_count, fix=True):
    from openpyxl import load_workbook

    workbook_path = folder / f"{folder.name}.xlsx"
    if not workbook_path.exists():
        return "", "", "否", "表格不存在"

    workbook = load_workbook(workbook_path)
    sheet = workbook["Sheet1"] if "Sheet1" in workbook.sheetnames else workbook.active
    part_count = sheet["C2"].value
    current_joint_count = sheet["D2"].value
    modified = "否"

    if as_int(current_joint_count) != as_int(usd_joint_count):
        current_joint_count = usd_joint_count
        if fix:
            sheet["D2"].value = usd_joint_count
            workbook.save(workbook_path)
            modified = "是"
        else:
            modified = "待修改"

    workbook.close()
    return part_count, current_joint_count, modified, "OK"


def main():
    setup_usd_runtime()
    args = parse_args()
    source = Path(args.source).expanduser().resolve()
    output = Path(args.output).expanduser().resolve()

    if not source.exists():
        raise SystemExit(f"Source does not exist: {source}")

    usd_files = collect_usd_files(source)
    output.parent.mkdir(parents=True, exist_ok=True)

    rows = [
        [
            "文件夹",
            "USD文件",
            "相对路径",
            "根Xform名称",
            "根Xform路径",
            "根Xform是否有ArticulationRoot",
            "ArticulationRoot路径",
            "ArticulationRoot检查结果",
            "Xform数量",
            "group* Xform数量",
            "非group*且非根Xform数量",
            "USD关节数量",
            "表格零件数量",
            "表格当前关节数量",
            "是否修改表格",
            "表格读取状态",
            "USD读取状态",
        ]
    ]
    ok = 0
    failed = 0
    modified_count = 0

    for usd_file in usd_files:
        folder = usd_file.parent
        folder_name = folder.name

        try:
            usd_info = inspect_usd(usd_file)
            usd_status = "OK"
            ok += 1
        except Exception as exc:
            usd_info = {
                "root_name": "",
                "root_path": "",
                "root_has_articulation": "",
                "articulation_paths": "",
                "articulation_status": "",
                "xform_total": "",
                "group_named": "",
                "non_group_non_root": "",
                "joint_total": "",
            }
            usd_status = f"ERROR: {exc}"
            failed += 1

        part_count, current_joint_count, modified, workbook_status = read_and_fix_delivery_counts(
            folder,
            usd_info["joint_total"],
            fix=not args.no_fix and usd_status == "OK",
        )
        if modified == "是":
            modified_count += 1

        relative_path = str(usd_file.relative_to(source)) if source.is_dir() else usd_file.name
        rows.append(
            [
                folder_name,
                usd_file.name,
                relative_path,
                usd_info["root_name"],
                usd_info["root_path"],
                usd_info["root_has_articulation"],
                usd_info["articulation_paths"],
                usd_info["articulation_status"],
                usd_info["xform_total"],
                usd_info["group_named"],
                usd_info["non_group_non_root"],
                usd_info["joint_total"],
                part_count,
                current_joint_count,
                modified,
                workbook_status,
                usd_status,
            ]
        )

    with output.open("w", newline="", encoding="utf-8-sig") as csv_file:
        writer = csv.writer(csv_file)
        writer.writerows(rows)

    missing_count = sum(1 for row in rows[1:] if row[7] != "OK")
    print(f"USD files: {len(usd_files)}")
    print(f"Counted: {ok}")
    print(f"Failed: {failed}")
    print(f"Modified workbooks: {modified_count}")
    print(f"ArticulationRoot issues: {missing_count}")
    print(f"Output CSV: {output}")


if __name__ == "__main__":
    main()
